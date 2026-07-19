from collections.abc import Sequence
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.engines.documents.branding import ClubBrandingDTO
from app.engines.documents.branding_render import (
    apply_excel_branding,
    apply_excel_table_style,
    apply_excel_title_style,
)
from app.engines.documents.consolidated_dto import ConsolidatedProjectDocumentDTO
from app.engines.documents.dto import GeneratedDocument

_GENERATION_MODE_LABELS = {
    "club_only": "Только клубные рецепты",
    "club_and_personal": "Клубные и личные рецепты",
    "personal_preferred": "Личные рецепты в приоритете",
}
_STATUS_LABELS = {
    "draft": "Черновик",
    "prepared": "Подготовлен",
    "ready": "Готов",
}


class ConsolidatedExcelDocumentGenerator:
    """Generate the complete Russian workbook defined by PRODUCT_SPEC."""

    @staticmethod
    def _prepare_sheet(
        workbook: Workbook,
        sheet: Worksheet,
        *,
        title: str,
        headers: Sequence[str] | None,
        branding: ClubBrandingDTO | None,
    ) -> int | None:
        apply_excel_branding(workbook, sheet, branding)
        sheet.append([title])
        apply_excel_title_style(
            sheet,
            row=1,
            last_column=max(len(headers or ()), 2),
            branding=branding,
        )
        if headers is None:
            return None
        sheet.append([])
        sheet.append(list(headers))
        return 3

    @staticmethod
    def _finish_table(
        sheet: Worksheet,
        *,
        header_row: int,
        last_column: int,
        branding: ClubBrandingDTO | None,
        widths: Sequence[float],
    ) -> None:
        apply_excel_table_style(
            sheet,
            header_row=header_row,
            last_row=sheet.max_row,
            last_column=last_column,
            branding=branding,
        )
        sheet.freeze_panes = f"A{header_row + 1}"
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.auto_filter.ref = (
            f"A{header_row}:{chr(64 + last_column)}{sheet.max_row}"
        )

    @staticmethod
    def _append_rows(sheet: Worksheet, rows: Sequence[Sequence[Any]]) -> None:
        if rows:
            for row in rows:
                sheet.append(list(row))
        else:
            sheet.append(["Нет данных"])

    def generate(
        self,
        document: ConsolidatedProjectDocumentDTO,
        branding: ClubBrandingDTO | None = None,
    ) -> GeneratedDocument:
        workbook = Workbook()
        trip_sheet = workbook.active
        trip_sheet.title = "Поход"
        self._prepare_sheet(
            workbook,
            trip_sheet,
            title="Параметры похода",
            headers=None,
            branding=branding,
        )
        trip_rows = [
            ("Название", document.summary.project_name),
            ("Идентификатор", document.summary.project_id),
            ("Участников", document.summary.participants),
            ("Дней", document.summary.days),
            ("Дата начала", document.summary.start_date or "Не указана"),
            ("Первый приём пищи", document.summary.first_meal or "Не указан"),
            ("Последний приём пищи", document.summary.last_meal or "Не указан"),
            (
                "Режим рецептов",
                _GENERATION_MODE_LABELS.get(
                    document.summary.recipe_generation_mode,
                    document.summary.recipe_generation_mode,
                ),
            ),
            (
                "Статус",
                _STATUS_LABELS.get(document.summary.status, document.summary.status),
            ),
            (
                "Ответственный за закупку",
                document.responsible_person or "Не указан",
            ),
        ]
        for label, value in trip_rows:
            trip_sheet.append([label, value])
        trip_sheet.append([])
        trip_sheet.append(["Предупреждения"])
        if document.warnings:
            for warning in document.warnings:
                trip_sheet.append([warning])
        else:
            trip_sheet.append(["Нет"])
        trip_sheet.append([])
        trip_sheet.append(["Комментарии"])
        if document.comments:
            for comment in document.comments:
                trip_sheet.append([comment])
        else:
            trip_sheet.append(["Нет"])
        trip_sheet.column_dimensions["A"].width = 36
        trip_sheet.column_dimensions["B"].width = 68
        trip_sheet.freeze_panes = "A2"

        menu_sheet = workbook.create_sheet("Меню")
        menu_header = self._prepare_sheet(
            workbook,
            menu_sheet,
            title="Меню по дням",
            headers=[
                "День",
                "Приём пищи",
                "Блюдо",
                "Рецепт",
                "Источник",
            ],
            branding=branding,
        )
        assert menu_header is not None
        self._append_rows(
            menu_sheet,
            [
                [
                    row.day_number,
                    row.meal_name,
                    row.dish_name,
                    row.recipe_name,
                    "Вручную" if row.is_manually_edited else "Автоматически",
                ]
                for row in document.menu_rows
            ],
        )
        self._finish_table(
            menu_sheet,
            header_row=menu_header,
            last_column=5,
            branding=branding,
            widths=[10, 22, 36, 40, 18],
        )

        loadout_sheet = workbook.create_sheet("Раскладка")
        loadout_header = self._prepare_sheet(
            workbook,
            loadout_sheet,
            title="Продуктовая раскладка",
            headers=["Продукт", "Категория", "Количество", "Единица"],
            branding=branding,
        )
        assert loadout_header is not None
        self._append_rows(
            loadout_sheet,
            [
                [
                    row.product_name,
                    row.category,
                    row.required_quantity,
                    row.required_unit,
                ]
                for row in document.loadout_rows
            ],
        )
        self._finish_table(
            loadout_sheet,
            header_row=loadout_header,
            last_column=4,
            branding=branding,
            widths=[38, 24, 16, 16],
        )

        purchase_sheet = workbook.create_sheet("Закупка")
        purchase_header = self._prepare_sheet(
            workbook,
            purchase_sheet,
            title="Закупка и упаковка",
            headers=[
                "Продукт",
                "Категория",
                "Нужно",
                "Единица",
                "Упаковка",
                "Ед. упаковки",
                "Упаковок",
                "Купить",
                "Остаток",
                "Куплено",
                "Статус",
                "Комментарий",
            ],
            branding=branding,
        )
        assert purchase_header is not None
        self._append_rows(
            purchase_sheet,
            [
                [
                    row.product_name,
                    row.category,
                    row.required_quantity,
                    row.required_unit,
                    row.package_size,
                    row.package_unit,
                    row.packages_count,
                    row.purchase_quantity,
                    row.surplus_quantity,
                    row.purchased_quantity,
                    "Куплено" if row.is_checked else "Не куплено",
                    row.comment or "",
                ]
                for row in document.purchase_rows
            ],
        )
        self._finish_table(
            purchase_sheet,
            header_row=purchase_header,
            last_column=12,
            branding=branding,
            widths=[34, 20, 13, 14, 13, 16, 12, 13, 13, 13, 14, 36],
        )

        equipment_sheet = workbook.create_sheet("Оборудование")
        equipment_header = self._prepare_sheet(
            workbook,
            equipment_sheet,
            title="Оборудование",
            headers=["Оборудование", "Требуется", "Расчёт", "Источник"],
            branding=branding,
        )
        assert equipment_header is not None
        self._append_rows(
            equipment_sheet,
            [
                [
                    row.equipment_name,
                    row.required_quantity,
                    row.calculated_quantity,
                    row.source,
                ]
                for row in document.equipment_rows
            ],
        )
        self._finish_table(
            equipment_sheet,
            header_row=equipment_header,
            last_column=4,
            branding=branding,
            widths=[44, 14, 14, 24],
        )

        workbook.properties.title = (
            f"Документы похода — {document.summary.project_name}"
        )
        buffer = BytesIO()
        workbook.save(buffer)
        return GeneratedDocument(
            filename=f"tourhub_project_{document.summary.project_id}_complete.xlsx",
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            generated_at=datetime.now(UTC),
            content=buffer.getvalue(),
        )
