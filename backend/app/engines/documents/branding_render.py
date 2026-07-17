from io import BytesIO
from typing import Any

from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image as PDFImage
from reportlab.platypus import Paragraph, Spacer

from app.engines.documents.branding import ClubBrandingDTO


def _scaled_pdf_image(image_bytes: bytes, *, max_width: float, max_height: float) -> PDFImage:
    image_data = BytesIO(image_bytes)
    width, height = ImageReader(image_data).getSize()
    scale = min(max_width / width, max_height / height)
    image_data.seek(0)
    return PDFImage(image_data, width=width * scale, height=height * scale)


def pdf_branding_flowables(
    styles: Any,
    branding: ClubBrandingDTO | None,
) -> list[Any]:
    if branding is None:
        return []

    flowables: list[Any] = []
    if branding.title_background_bytes is not None:
        flowables.append(
            _scaled_pdf_image(
                branding.title_background_bytes,
                max_width=170 * mm,
                max_height=35 * mm,
            )
        )
        flowables.append(Spacer(1, 3 * mm))

    if branding.logo_bytes is not None:
        flowables.append(
            _scaled_pdf_image(
                branding.logo_bytes,
                max_width=28 * mm,
                max_height=18 * mm,
            )
        )
        flowables.append(Spacer(1, 2 * mm))

    heading_style = styles["Heading2"].clone("TourHubClubHeading")
    heading_style.textColor = colors.HexColor(branding.palette.heading_color)
    flowables.append(Paragraph(branding.club_name, heading_style))

    if branding.contact_lines:
        contact_style = styles["Normal"].clone("TourHubClubContacts")
        contact_style.fontSize = 8
        contact_style.leading = 10
        for line in branding.contact_lines:
            flowables.append(Paragraph(line, contact_style))

    flowables.append(Spacer(1, 3 * mm))
    return flowables


def pdf_table_style_commands(
    branding: ClubBrandingDTO | None,
    *,
    font_name: str,
) -> list[tuple[Any, ...]]:
    if branding is None:
        header_background = colors.lightgrey
        header_text = colors.black
        border = colors.black
        compact = False
    else:
        header_background = colors.HexColor(branding.palette.table_header_background)
        header_text = colors.HexColor(branding.palette.table_header_text)
        border = colors.HexColor(branding.palette.table_border_color)
        compact = branding.table_density == "compact"

    font_size = 8 if compact else 9
    vertical_padding = 3 if compact else 5
    return [
        ("FONT", (0, 0), (-1, -1), font_name, font_size),
        ("BACKGROUND", (0, 0), (-1, 0), header_background),
        ("TEXTCOLOR", (0, 0), (-1, 0), header_text),
        ("GRID", (0, 0), (-1, -1), 0.5, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), vertical_padding),
        ("BOTTOMPADDING", (0, 0), (-1, -1), vertical_padding),
    ]


def apply_pdf_heading_styles(styles: Any, branding: ClubBrandingDTO | None) -> None:
    if branding is None:
        return
    heading_color = colors.HexColor(branding.palette.heading_color)
    styles["Title"].textColor = heading_color
    styles["Heading1"].textColor = heading_color
    styles["Heading2"].textColor = heading_color


def _excel_color(value: str) -> str:
    return value.removeprefix("#").upper()


def _add_excel_image(
    sheet: Any,
    image_bytes: bytes,
    anchor: str,
    *,
    max_width: int,
    max_height: int,
) -> None:
    image = ExcelImage(BytesIO(image_bytes))
    scale = min(max_width / image.width, max_height / image.height, 1)
    image.width *= scale
    image.height *= scale
    sheet.add_image(image, anchor)


def apply_excel_branding(
    workbook: Any,
    sheet: Any,
    branding: ClubBrandingDTO | None,
) -> None:
    if branding is None:
        return

    workbook.properties.creator = branding.club_name
    sheet.oddHeader.center.text = branding.club_name
    footer = branding.footer_text or f"Сформировано для {branding.club_name} в TourHub"
    sheet.oddFooter.left.text = footer
    sheet.oddFooter.right.text = "Страница &P из &N"

    if branding.logo_bytes is not None:
        _add_excel_image(sheet, branding.logo_bytes, "F1", max_width=120, max_height=60)
    if branding.title_background_bytes is not None:
        _add_excel_image(
            sheet,
            branding.title_background_bytes,
            "H1",
            max_width=220,
            max_height=90,
        )

    for index, line in enumerate(branding.contact_lines, start=4):
        cell = sheet.cell(row=index, column=6, value=line)
        cell.font = Font(size=9, color=_excel_color(branding.palette.heading_color))


def apply_excel_title_style(
    sheet: Any,
    *,
    row: int,
    last_column: int,
    branding: ClubBrandingDTO | None,
) -> None:
    if branding is None:
        return
    fill = PatternFill(
        "solid",
        fgColor=_excel_color(branding.palette.title_background_color),
    )
    for column in range(1, last_column + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = fill
        cell.font = Font(
            bold=True,
            size=14,
            color=_excel_color(branding.palette.heading_color),
        )
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 26


def apply_excel_table_style(
    sheet: Any,
    *,
    header_row: int,
    last_row: int,
    last_column: int,
    branding: ClubBrandingDTO | None,
) -> None:
    if branding is None:
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
        return

    header_fill = PatternFill(
        "solid",
        fgColor=_excel_color(branding.palette.table_header_background),
    )
    border_side = Side(
        style="thin",
        color=_excel_color(branding.palette.table_border_color),
    )
    border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )
    row_height = 18 if branding.table_density == "compact" else 24

    for row in range(header_row, last_row + 1):
        sheet.row_dimensions[row].height = row_height
        for column in range(1, last_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row == header_row:
                cell.fill = header_fill
                cell.font = Font(
                    bold=True,
                    color=_excel_color(branding.palette.table_header_text),
                )
