from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from app.core.database import get_db
from app.main import app
from app.models.equipment_list import EquipmentListORM
from app.models.equipment_list_item import EquipmentListItemORM
from app.models.meal_plan import MealPlanORM

from tests.api.test_project_documents_success_api import (
    create_project_with_purchase_list,
    override_test_db,
)


def _add_equipment_list(db_session, project_id: int) -> None:
    meal_plan = MealPlanORM(
        id="meal-plan-doc-test",
        project_id=project_id,
        name="Document meal plan",
        participants=4,
        days_count=3,
        warnings=["Проверьте разнообразие меню."],
    )
    equipment_list = EquipmentListORM(
        id="equipment-doc-test",
        project_id=project_id,
        meal_plan_id=meal_plan.id,
        status="prepared",
    )
    equipment_list.items.append(
        EquipmentListItemORM(
            id="equipment-item-doc-test",
            equipment_name="Котёл",
            required_quantity=2,
            calculated_quantity=2,
            is_manual=False,
            is_removed=False,
        )
    )
    db_session.add_all([meal_plan, equipment_list])
    db_session.commit()


def test_project_consolidated_documents(client, db_session):
    app.dependency_overrides[get_db] = override_test_db(db_session)

    project_id = create_project_with_purchase_list(db_session)
    _add_equipment_list(db_session, project_id)

    pdf_response = client.get(
        f"/api/v1/projects/{project_id}/documents/consolidated/pdf"
    )
    excel_response = client.get(
        f"/api/v1/projects/{project_id}/documents/consolidated/excel"
    )

    assert pdf_response.status_code == 200
    assert pdf_response.headers["content-type"] == "application/pdf"
    assert pdf_response.content.startswith(b"%PDF")

    assert excel_response.status_code == 200
    assert excel_response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(excel_response.content), data_only=True)
    assert workbook.sheetnames == [
        "Поход",
        "Меню",
        "Раскладка",
        "Закупка",
        "Оборудование",
    ]
    assert workbook["Поход"]["B2"].value == "Document Test Project"
    assert workbook["Раскладка"]["A4"].value == "Rice Document Test"

    app.dependency_overrides.clear()


def test_project_document_package(client, db_session):
    app.dependency_overrides[get_db] = override_test_db(db_session)

    project_id = create_project_with_purchase_list(db_session)
    _add_equipment_list(db_session, project_id)

    response = client.get(f"/api/v1/projects/{project_id}/documents/package")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"

    archive = ZipFile(BytesIO(response.content))
    assert set(archive.namelist()) == {
        f"tourhub_project_{project_id}_complete.pdf",
        f"tourhub_project_{project_id}_complete.xlsx",
        "purchase_list.pdf",
        "purchase_list.xlsx",
        "purchase_list.txt",
        "equipment_list.pdf",
        "equipment_list.xlsx",
    }

    app.dependency_overrides.clear()


def test_project_document_package_not_found(client, db_session):
    app.dependency_overrides[get_db] = override_test_db(db_session)

    response = client.get("/api/v1/projects/999999/documents/package")

    assert response.status_code == 404
    assert response.json()["error"] == "Project not found"

    app.dependency_overrides.clear()
