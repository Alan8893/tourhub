from io import BytesIO

from openpyxl import load_workbook

from app.engines.documents.branding import ClubBrandingDTO
from app.engines.documents.consolidated_dto import (
    ConsolidatedProjectDocumentDTO,
    EquipmentRowDTO,
    LoadoutRowDTO,
    MenuRowDTO,
    ProjectSummaryDTO,
    PurchaseRowDTO,
)
from app.engines.documents.consolidated_excel import (
    ConsolidatedExcelDocumentGenerator,
)
from app.engines.documents.consolidated_pdf import ConsolidatedPDFDocumentGenerator


def _document() -> ConsolidatedProjectDocumentDTO:
    return ConsolidatedProjectDocumentDTO(
        summary=ProjectSummaryDTO(
            project_id=42,
            project_name="Карельский маршрут",
            participants=8,
            days=4,
            start_date="2026-08-01",
            first_meal="dinner",
            last_meal="breakfast",
            recipe_generation_mode="club_and_personal",
            status="prepared",
        ),
        menu_rows=(
            MenuRowDTO(
                day_number=1,
                meal_type="dinner",
                meal_name="Ужин",
                dish_name="Плов",
                recipe_name="Плов клубный",
                is_manually_edited=False,
            ),
            MenuRowDTO(
                day_number=2,
                meal_type="breakfast",
                meal_name="Завтрак",
                dish_name="Каша",
                recipe_name="Каша инструктора",
                is_manually_edited=True,
            ),
        ),
        loadout_rows=(
            LoadoutRowDTO(
                product_name="Рис",
                category="Крупы",
                required_quantity=2400,
                required_unit="gram",
            ),
        ),
        purchase_rows=(
            PurchaseRowDTO(
                product_name="Рис",
                category="Крупы",
                required_quantity=2400,
                required_unit="gram",
                package_size=900,
                package_unit="gram",
                packages_count=3,
                purchase_quantity=2700,
                surplus_quantity=300,
                purchased_quantity=1800,
                is_checked=False,
                comment="Докупить одну упаковку",
            ),
        ),
        equipment_rows=(
            EquipmentRowDTO(
                equipment_name="Котёл 8 л",
                required_quantity=2,
                calculated_quantity=2,
                source="Расчёт",
            ),
        ),
        warnings=("Недостаточно разнообразных основных блюд.",),
        comments=("Проверить газ перед выездом.",),
        responsible_person="Иван Петров",
    )


def test_consolidated_pdf_contains_complete_project_document() -> None:
    generated = ConsolidatedPDFDocumentGenerator().generate(
        _document(),
        ClubBrandingDTO(club_name="Турклуб Север"),
    )

    assert generated.filename == "tourhub_project_42_complete.pdf"
    assert generated.content_type == "application/pdf"
    assert isinstance(generated.content, bytes)
    assert generated.content.startswith(b"%PDF")
    assert len(generated.content) > 2_000


def test_consolidated_excel_has_approved_russian_sheets_and_content() -> None:
    generated = ConsolidatedExcelDocumentGenerator().generate(
        _document(),
        ClubBrandingDTO(club_name="Турклуб Север"),
    )

    assert generated.filename == "tourhub_project_42_complete.xlsx"
    assert isinstance(generated.content, bytes)
    workbook = load_workbook(BytesIO(generated.content), data_only=True)

    assert workbook.sheetnames == [
        "Поход",
        "Меню",
        "Раскладка",
        "Закупка",
        "Оборудование",
    ]
    assert workbook["Поход"]["B2"].value == "Карельский маршрут"
    assert workbook["Меню"]["C4"].value == "Плов"
    assert workbook["Меню"]["D5"].value == "Каша инструктора"
    assert workbook["Раскладка"]["A4"].value == "Рис"
    assert workbook["Закупка"]["G4"].value == 3
    assert workbook["Закупка"]["L4"].value == "Докупить одну упаковку"
    assert workbook["Оборудование"]["A4"].value == "Котёл 8 л"
